import tkinter
import openpyxl
from tkinter import filedialog

root = tkinter.Tk()
root.withdraw()


def save_as(name):
    print("Выбери место сохранения...")
    directory = filedialog.asksaveasfilename(initialfile=name)
    return directory


class Table:
    data = []
    total = 0

    def __init__(self, path):
        self.wb = openpyxl.Workbook()
        self.path = path

    def collect_questions(self, data):
        self.data.extend(data)
        print(f"-- Собрано {len(data)} вопросов!")
        self.total += len(data)
        return len(data)

    def max_answer_count(self):
        max_count = 0
        for question in self.data:
            if max_count < len(question['answers']):
                max_count = len(question['answers'])

        print(f"-- Максимальное количество ответов: {max_count}")
        return max_count

    def generate_table(self):
        alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        max_count = self.max_answer_count()
        ws = self.wb.active

        ws[f'{alphabet[0]}1'] = "Question Text"
        ws[f'{alphabet[1]}1'] = "Question Type"
        for i in range(max_count):
            ws[f'{alphabet[2 + i]}1'] = f"Option {i + 1}"
        ws[f'{alphabet[2 + max_count]}1'] = "Correct Answer"
        ws[f'{alphabet[3 + max_count]}1'] = "Time in seconds"

        row = 2
        for question in self.data:
            ws[f'{alphabet[0]}{row}'] = question['question']
            if len(question['correctAnswer']) > 1:
                ws[f'{alphabet[1]}{row}'] = "Checkbox"
            else:
                ws[f'{alphabet[1]}{row}'] = "Multiple Choice"

            for current_option in range(max_count):
                if current_option < len(question['answers']):
                    ws[f'{alphabet[2 + current_option]}{row}'] = question['answers'][current_option]
                else:
                    ws[f'{alphabet[2 + current_option]}{row}'] = ""

            correct_answers = []
            for correct in question['correctAnswer']:
                correct_answers.append(question['answers'].index(correct) + 1)
            ws[f'{alphabet[2 + max_count]}{row}'] = str(", ".join(map(str, correct_answers)))

            ws[f'{alphabet[3 + max_count]}{row}'] = 20

            row += 1

        self.wb.save(self.path)
